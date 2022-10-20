from json import load
import os
import pickle
import sys
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

def mark_attendence_of_students(people_found):
    try:
        f = open("./important files/attendence.pickle", 'rb')
        attendence = pickle.load(f)
        f.close()
    except:
        print("Something went wrong!--1")
        sys.exit(1)
    
    for person in people_found:
        if person in attendence:
            attendence[person] += 1
    attendence['total'] += 1

    try:
        f = open("./important files/attendence.pickle", 'wb')
        pickle.dump(attendence, f)
        f.close()
    except:
        print("Something went wrong!--2")
        sys.exit(1)

    try : 
        wb = load_workbook("./Attendence_excel.xlsx")
        ws = wb['Attendence']
        new_col = attendence['total'] + 1
        col_char = get_column_letter(new_col)
        ws[col_char + '1'] = date.today().strftime("%m/%d/%Y")
        for row in range(1, 1+len(attendence.keys())):
            ws[col_char+str(row+1)].value = attendence[list(attendence.keys())[row-1]]
        wb.save("./Attendence_excel.xlsx")
    except Exception as e:
        print(e)
        print("Something went wrong!--3")
        sys.exit(1)

def make_attendence_dataset():
    dataset_dir = "./dataset"
    attendence = {}
    for person in os.listdir(dataset_dir):
        attendence[person] = 0
    attendence['total'] = 0
    try:
        f = open("./important files/attendence.pickle", 'wb')
        pickle. dump(attendence, f)
        f.close()
    except:
        print("Something went wrong!--4")
        sys.exit(1)

    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Attendence"
        ws['A1'] = "Names"
        for row in range(1, 1+len(attendence.keys())):
            ws['A'+str(row+1)].value = list(attendence.keys())[row-1]
        wb.save("./Attendence_excel.xlsx")
    except:
        print("Something went wrong!--6")
        sys.exit(1)

def read_attendence():
    try:
        f = open("./important files/attendence.pickle", 'rb')
        attendence = pickle.load(f)
        f.close()
    except:
        print("Something went wrong!--7")
        sys.exit(1)
    return attendence

if __name__ == '__main__':
    make_attendence_dataset()